# —— coding :utf-8 ——
# @time:    2020/9/14 17:50
# @IDE:     Test_HttprequestDemo
# @Author:  jsonJie
# @Email:   810030907@qq.com
# @File:    handle_excel2.py
from openpyxl import load_workbook


class HandleExcel:
    '''需要用到时候读取所有数据，对磁盘读写要求高点'''

    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.filename)[self.sheet_name]
        # 获取一个表单对象

    def get_data(self, i, j):
        '''根据传入坐标获取值'''
        return self.sheet_obj.cell(i, j).value


if __name__ == '__main__':
    res = HandleExcel('test_data.xlsx', 'login').get_data(1, 1)
    print(res)
