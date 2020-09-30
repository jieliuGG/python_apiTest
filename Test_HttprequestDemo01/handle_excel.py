# —— coding :utf-8 ——
# @time:    2020/9/14 16:42
# @IDE:     Test_HttprequestDemo
# @Author:  jsonJie
# @Email:   810030907@qq.com
# @File:    handle_excel.py

# openpyxl只支持.xlsx格式的excel
# from openpyxl import load_workbook
# # 1. 打开excel
# wb = load_workbook('test_data.xlsx')
# # 2. 定位表单
# sheet = wb['login']
# # 3. 定位单元格 行列值
# res = sheet.cell(1,1).vaule


from openpyxl import load_workbook


class HandleExcel:
    '''一次性读取所有数据，对内存要求高点'''

    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row + 1):
            sub_data = []
            sub_data['method'] = sheet.cell(i, 1).value
            sub_data['url'] = sheet.cell(i, 2).value
            sub_data['data'] = sheet.cell(i, 3).value
            sub_data['expected'] = sheet.cell(i, 4)
            test_data.append(sub_data)
        return test_data  # 返回获取到的数据

    @staticmethod
    def write_back(filename, sheet_name, row, rol, result):
        '''专门写回数据'''
        wb = load_workbook(filename)
        sheet = wb(sheet_name)
        sheet.cell(row, rol).value = result
        wb.save(filename)  # 保存结果


if __name__ == '__main__':
    print(HandleExcel('test_data.xlsx', 'login').get_data())
