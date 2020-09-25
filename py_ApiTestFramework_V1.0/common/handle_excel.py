# —— coding :utf-8 ——
# @time:    2020/9/23 21:12
# @IDE:     py_ApiTestFramework_V1.0
# @Author:  jsonJie
# @Email:   810030907@qq.com
# @File:    handle_excel.py
from openpyxl import load_workbook
from common.file_dir import test_data_path
from common.file_dir import case_config_path
from common.read_config import ReadConfig
from common.get_data import GetData


class HandleExcel:
    @classmethod
    def get_data(cls, file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig().read_config(case_config_path, 'MODE', 'mode'))
        tel = getattr(GetData, 'NoRegTel')  # 利用反射拿到数据
        test_data = []
        for key in mode:
            sheet = wb[key]
            if mode[key] == 'all':
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    row_data['test_id'] = sheet.cell(i, 1).value
                    row_data['url'] = sheet.cell(i, 2).value
                    if sheet.cell(i, 3).value.find('${tel}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${tel}', str(tel))
                    elif sheet.cell(i, 3).value.find('${normal_tel}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${normal_tel}',
                                                                          str(getattr(GetData, 'normal_tel')))
                    elif sheet.cell(i, 3).value.find('${admin_tel}') != -1:  # 不为空做替换
                        row_data['data'] = sheet.cell(i, 3).value.replace('${admin_tel}',
                                                                          str(getattr(GetData, 'admin_tel')))
                    elif sheet.cell(i, 3).value.find('${loan_member_id}') != -1:  # 不为空做替换
                        row_data['data'] = sheet.cell(i, 3).value.replace('${loan_member_id}',
                                                                          str(getattr(GetData, 'loan_member_id')))
                    elif sheet.cell(i, 3).value.find('${memberID}') != -1:  # 不为空做替换
                        row_data['data'] = sheet.cell(i, 3).value.replace('${memberID}',
                                                                          str(getattr(GetData, 'memberID')))
                    else:
                        row_data = sheet.cell(i, 3).value
                    if sheet.cell(i, 4).value.find('${normal_tel}') != -1:
                        row_data['check_sql'] = sheet.cell(i, 4).value.replace('${normal_tel}',
                                                                               str(getattr(GetData, 'normal_tel')))
                    else:
                        row_data['check_sql'] = sheet.cell(i, 4).value
                    row_data['title'] = sheet.cell(i, 5).value
                    row_data['http_method'] = sheet.cell(i, 6).value
                    row_data['excepted'] = sheet.cell(i, 7).value
                    sheet['sheet_name'] = key
                    test_data.append(row_data)
                    cls.update_tel(getattr(GetData, 'NoRegTel'), file_name, 'init')  # 利用类方法更新手机号
            else:
                for test_id in mode[key]:  # test_id为用例id
                    row_data = {}
                    row_data['test_id'] = sheet.cell(test_id + 1, 1).value
                    row_data['url'] = sheet.cell(test_id + 1, 2).value
                    if sheet.cell(test_id + 1, 3).value.find('${tel}') != -1:
                        row_data['data'] = sheet.cell(test_id + 1, 3).value.replace('${tel}', str(tel))
                    elif sheet.cell(test_id + 1, 3).value.find('${normal_tel}') != -1:
                        row_data['data'] = sheet.cell(test_id + 1, 3).value.replace('${normal_tel}',
                                                                                    str(getattr(GetData, 'normal_tel')))
                    elif sheet.cell(test_id + 1, 3).value.find('${admin_tel}') != -1:  # 不为空做替换
                        row_data['data'] = sheet.cell(test_id + 1, 3).value.replace('${admin_tel}',
                                                                                    str(getattr(GetData, 'admin_tel')))
                    elif sheet.cell(test_id + 1, 3).value.find('${loan_member_id}') != -1:  # 不为空做替换
                        row_data['data'] = sheet.cell(test_id + 1, 3).value.replace('${loan_member_id}',
                                                                                    str(getattr(GetData,
                                                                                                'loan_member_id')))
                    elif sheet.cell(test_id + 1, 3).value.find('${memberID}') != -1:  # 不为空做替换
                        row_data['data'] = sheet.cell(test_id + 1, 3).value.replace('${memberID}',
                                                                                    str(getattr(GetData, 'memberID')))
                    else:
                        row_data = sheet.cell(test_id + 1, 3).value

                    if sheet.cell(test_id + 1, 4).value.find('${normal_tel}') != -1:
                        row_data['check_sql'] = sheet.cell(test_id + 1, 4).value.replace('${normal_tel}',
                                                                                    str(getattr(GetData,'normal_tel')))
                    else:
                        row_data['check_sql'] = sheet.cell(test_id + 1, 4).value

                    row_data['title'] = sheet.cell(test_id + 1, 5).value
                    row_data['http_method'] = sheet.cell(test_id + 1, 6).value
                    row_data['excepted'] = sheet.cell(test_id + 1, 7).value
                    sheet['sheet_name'] = key
                    test_data.append(row_data)
                return test_data

    @staticmethod
    def write_back(file_name, sheet_name, row, rol, result):
        """
        专门写回数据
        :param file_name: 文本名称
        :param sheet_name: 工作簿名称
        :param row: 行
        :param rol: 列
        :param result: 写回到工作簿的结果
        :return:
        """
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(row, rol).value = result
        wb.save(file_name)

    @classmethod
    def update_tel(cls, tel, filename, sheet_name):
        '''更新excel里的手机号数据'''
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
        sheet.cell(2, 1).value = tel
        wb.save(filename)


if __name__ == '__main__':
    test_data = HandleExcel().get_data(test_data_path)
    print(test_data)
