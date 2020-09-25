# —— coding :utf-8 ——
# @time:    2020/9/23 22:01
# @IDE:     py_ApiTestFramework_V1.0
# @Author:  jsonJie
# @Email:   810030907@qq.com
# @File:    handle_baseexcel.py
from openpyxl import load_workbook
from common.file_dir import test_data_path
class HandleExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(2,sheet.max_row+1):
            sub_data = {}
            sub_data['test_id'] = sheet.cell(i,1).value
            sub_data['url'] = sheet.cell(i,2).value
            sub_data['title'] = sheet.cell(i,3).value
            sub_data['http_method'] =sheet.cell(i,4).value
            sub_data['expected'] = sheet.cell(i,5).value
            sub_data['result'] = sheet.cell(i,6).value
            sub_data['TestResult'] = sheet.cell(i,7).value
            test_data.append(sub_data)
            return test_data    # 返回获取到的数据

    @staticmethod
    def write_back(file_name,sheet_name,row,rol,result):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.chell(row,rol).value = result
        wb.save(file_name) # 保存结果

    @staticmethod
    def update_tel(file_name,sheet_name,tel):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.chell(2,1).value = tel
        wb.save(file_name)
if __name__ == '__main__':
    print(HandleExcel(test_data_path,'login').get_data())

