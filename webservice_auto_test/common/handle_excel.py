# —— coding :utf-8 ——
# @time:    2020/9/15 0:44
# @IDE:     webservice_auto_test
# @Author:  jsonJie
# @Email:   810030907@qq.com
# @File:    handle_excel.py
from gettext import find
import pandas as pd

from openpyxl import load_workbook
from common.handle_regx import HandleRegx

class HandleExcel:
    '''一次性读取所有数据，对内存要求高点'''


    def handle_excel(self,file_name,sheet_name):
        wb = load_workbook(file_name,sheet_name)
        sheet = wb[sheet_name]
        test_data = []
        for i in range(2,sheet.max_row+1):
            row_data = {}
            row_data['case_id'] = sheet_name.cell(i,1).value
            row_data['title'] = sheet_name.cell(i,2).value
            row_data['param'] = HandleRegx().handle_Reg(sheet.cell(i,3).value,file_name)
            row_data['expected'] = sheet_name.cell(i,4).value
            test_data.append(row_data)
        return test_data
    @staticmethod
    def write_back(filename, sheet_name, row, rol, result):
        '''专门写回数据'''
        wb = load_workbook(filename)
        sheet = wb(sheet_name)
        sheet.cell(row, rol).value = result
        wb.save(filename)  # 保存结果


if __name__ == '__main__':
    test_data= HandleExcel('test_data.xlsx', 'login').get_data()
    print(test_data)
